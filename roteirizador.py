"""
=============================================================
  ROTEIRIZADOR — Sistema de Itinerário
=============================================================
  Como usar:
    1. Coloque este arquivo na mesma pasta que o Excel
    2. Ajuste ARQUIVO_EXCEL e CEP_ORIGEM abaixo
    3. Execute: python roteirizador.py

  Dependências (instalar uma vez):
    pip install pandas openpyxl requests
=============================================================
"""

# ─────────────────────────────────────────────────────────────
#  IMPORTS
# ─────────────────────────────────────────────────────────────

import os, sys, json, math, time, threading, re
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm

# ─────────────────────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────────────────────

ARQUIVO_EXCEL = ""
CEP_ORIGEM           = "13347-402"   # <- CEP da sua empresa
VELOCIDADE_MEDIA_KMH = 35
BUFFER_ENTRE_PARADAS = 10            # minutos de folga entre paradas

# Tempo padrão de serviço por tipo (usado quando campo estiver vazio no pedido)
TEMPO_PADRAO_POR_TIPO = {
    "manutenção": 60,
    "manutencao": 60,
    "entrega":    30,
    "itens":      30,
}
TEMPO_PADRAO_GERAL = 30
BASE_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))

PASTA_PDF = os.path.join(BASE_DIR, "PDFs")

os.makedirs(PASTA_PDF, exist_ok=True)


# ─────────────────────────────────────────────────────────────
#  CACHE DE CEPs
# ─────────────────────────────────────────────────────────────

CACHE_FILE = "cep_cache.json"

def carregar_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def salvar_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

cep_cache = carregar_cache()

# ─────────────────────────────────────────────────────────────
#  FUNÇÕES DE CÁLCULO
# ─────────────────────────────────────────────────────────────

def limpar_cep(cep):
    return str(cep).replace("-", "").replace(".", "").strip().zfill(8)


def buscar_coordenadas(cep, log_fn=None):
    cep_limpo = limpar_cep(cep)
    if cep_limpo in cep_cache:
        d = cep_cache[cep_limpo]
        if d.get("lat") and d.get("lon"):
            return d["lat"], d["lon"]
    if log_fn:
        log_fn(f"Buscando CEP {cep_limpo}...")
    try:
        r = requests.get(
            f"https://viacep.com.br/ws/{cep_limpo}/json/",
            timeout=10
        )
        d = r.json()
        if d.get("erro"):
            if log_fn:
                log_fn(f"⚠ CEP {cep_limpo} não encontrado.")
            return None, None
    except Exception as e:
        if log_fn:
            log_fn(f"⚠ Erro ViaCEP: {e}")
        return None, None

    query = (
        f"{d.get('logradouro', '')}, "
        f"{d.get('localidade', '')}, "
        f"{d.get('uf', '')}, Brasil"
    )
    try:
        time.sleep(1)
        r2 = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"format": "json", "q": query, "limit": 1},
            headers={"User-Agent": "Roteirizador/1.0"},
            timeout=10,
        )
        nd = r2.json()
        if nd:
            lat = float(nd[0]["lat"])
            lon = float(nd[0]["lon"])
            cep_cache[cep_limpo] = {
                "logradouro": d.get("logradouro", ""),
                "cidade":     d.get("localidade", ""),
                "uf":         d.get("uf", ""),
                "lat": lat,
                "lon": lon,
            }
            salvar_cache(cep_cache)
            return lat, lon
    except Exception as e:
        if log_fn:
            log_fn(f"⚠ Erro Nominatim: {e}")
    return None, None


def haversine(lat1, lon1, lat2, lon2):
    if None in (lat1, lon1, lat2, lon2):
        return 0.0
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1))
         * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def hora_para_minutos(valor):
    """
    Aceita: fração decimal do Excel (0.291=07:00), "HH:MM",
    "HH:MM:SS", pandas Timestamp, datetime.time, int/str.
    """
    try:
        if valor is None or str(valor).strip() in ("", "nan", "NaT"):
            return 0
        if hasattr(valor, "hour"):
            return valor.hour * 60 + valor.minute
        s = str(valor).strip()
        try:
            f = float(s)
            if 0.0 <= f < 1.0:
                return round(f * 24 * 60)
        except ValueError:
            pass
        partes = s.split(":")
        if len(partes) >= 2:
            return int(partes[0]) * 60 + int(partes[1])
        return int(s) * 60
    except Exception as e:
        print(f"AVISO hora_para_minutos({valor!r}): {e}")
        return 0

def minutos_para_hora(minutos):
    minutos = max(0, int(minutos))
    return f"{minutos // 60:02d}:{minutos % 60:02d}"


def somar_minutos(hora_str, delta):
    return minutos_para_hora(hora_para_minutos(hora_str) + delta)


def tempo_viagem_min(dist_km):
    if dist_km == 0:
        return 0
    return (dist_km / VELOCIDADE_MEDIA_KMH) * 60


def dia_semana_ptbr(data_str):
    dias = {
        0: "Segunda",
        1: "Terça",
        2: "Quarta",
        3: "Quinta",
        4: "Sexta",
        5: "Sábado",
        6: "Domingo"
    }

    data = pd.to_datetime(data_str)

    return dias[data.weekday()]

def calcular_score(p, lat_cur, lon_cur, hora_cur_min):
    """
    Score menor = melhor candidato para ser visitado agora.

    Simula a chegada real e penaliza:
      - tempo de espera (chegou antes da abertura)
      - atraso           (chegou depois do fechamento)
      - distância        (quanto mais longe, pior)
      - retirada externa (sempre por último — penalidade máxima)
      - urgência         (urgente sobe na fila — bônus negativo)
    """
    dist        = haversine(lat_cur, lon_cur, p.get("lat"), p.get("lon"))
    mins_viagem = tempo_viagem_min(dist)
    chegada     = hora_cur_min + mins_viagem

    abertura    = hora_para_minutos(p["h_entrada"])
    fechamento  = hora_para_minutos(p["h_saida"])

    # Tempo perdido esperando a abertura
    espera = max(0, abertura - chegada)

    # Penalidade por chegar depois do fechamento (inviável)
    atraso = 10000 if chegada > fechamento else 0

    # Penalidade de retirada externa (garante que vai por último)
    pen_retirada = 100000 if p.get("retirada_ext") else 0

    # Bônus de urgência (score negativo = sobe na fila)
    bonus_urgencia = -50000 if p.get("prioridade") == "urgente" else 0

    return dist + espera + atraso + pen_retirada + bonus_urgencia


def montar_rota(orig_lat, orig_lon, hora_ini_min, paradas):
    """
    Constrói a rota passo a passo.
    A cada passo escolhe a parada com menor score considerando
    distância + espera + janela de horário + urgência + retirada.
    """
    nao_visit  = list(paradas)
    rota       = []
    lat_cur    = orig_lat
    lon_cur    = orig_lon
    hora_cur   = hora_ini_min

    while nao_visit:
        # Ordena candidatos pelo score a partir do ponto atual
        candidatos = sorted(
            nao_visit,
            key=lambda p: calcular_score(p, lat_cur, lon_cur, hora_cur)
        )
        melhor = candidatos[0]

        # Registra distância desde o ponto anterior
        dist = haversine(lat_cur, lon_cur, melhor.get("lat"), melhor.get("lon"))
        melhor["dist_anterior_km"] = round(dist, 2)
        rota.append(melhor)

        # Avança posição e hora para o próximo passo
        lat_cur  = melhor.get("lat") or lat_cur
        lon_cur  = melhor.get("lon") or lon_cur

        # Hora real de saída deste cliente (inclui espera se necessário)
        chegada      = hora_cur + tempo_viagem_min(dist)
        abertura     = hora_para_minutos(melhor["h_entrada"])
        chegada_efet = max(chegada, abertura)
        hora_cur     = chegada_efet + melhor["tempo_est"] + BUFFER_ENTRE_PARADAS

        nao_visit = [p for p in nao_visit if p is not melhor]

    return rota 

def selecionar_planilha():

    caminho = filedialog.askopenfilename(
        title="Selecione a planilha",
        filetypes=[
            ("Arquivos Excel", "*.xlsx *.xls")
        ]
    )

    return caminho

# ─────────────────────────────────────────────────────────────
#  LEITURA DO EXCEL
# ─────────────────────────────────────────────────────────────

def ler_excel():
    abas = ["Clientes", "Horarios_Clientes", "Pedidos", "Funcionarios"]
    dfs = {}
    try:
        for aba in abas:
            df = pd.read_excel(ARQUIVO_EXCEL, sheet_name=aba, dtype=str).fillna("")
            df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
                # Ajustar apenas a data_trabalho da aba Pedidos
            if aba == "Pedidos" and "data_trabalho" in df.columns:

                df["data_trabalho"] = pd.to_datetime(
                    df["data_trabalho"],
                    errors="coerce"
                ).dt.strftime("%Y-%m-%d")

            dfs[aba] = df
    except FileNotFoundError:
        messagebox.showerror(
            "Arquivo não encontrado",
            f"'{ARQUIVO_EXCEL}' não encontrado.\n"
            "Coloque o script na mesma pasta do Excel."
        )
        sys.exit(1)
    except Exception as e:
        messagebox.showerror("Erro ao ler Excel", str(e))
        sys.exit(1)
    return dfs["Clientes"], dfs["Horarios_Clientes"], dfs["Pedidos"], dfs["Funcionarios"]

# ─────────────────────────────────────────────────────────────
#  CÁLCULO DA ROTA
# ─────────────────────────────────────────────────────────────

def calcular_rota(
    data,
    nome_func,
    jornada_entrada,
    jornada_saida,
    almoco_inicio,
    almoco_fim,
    clientes,
    horarios,
    pedidos,
    log_fn=None
):

    dia_semana    = dia_semana_ptbr(data)
    jornada_total = hora_para_minutos(jornada_saida) - hora_para_minutos(jornada_entrada)

    # ── 1. Filtrar pedidos ───────────────────────────────────
    status_exc = ["concluido", "concluído", "cancelado"]
    mask = (
        (pedidos["data_trabalho"].str.strip() == data) &
        (pedidos["funcionario"].str.strip().str.lower() == nome_func.strip().lower()) &
        (~pedidos["status"].str.lower().str.strip().isin(status_exc))
    )
    pedidos_dia = pedidos[mask].copy()

    if pedidos_dia.empty:
        return [], None, f"Nenhum pedido pendente para {nome_func} em {data}."

    # ── 2. Montar paradas ────────────────────────────────────
    paradas = []
    for _, ped in pedidos_dia.iterrows():
        fl_ret   = str(ped.get("fl_retirada",  "")).strip().lower()
        tipo_ret = str(ped.get("tipo_retirada", "")).strip().lower()

        # Retirada interna: cliente vem até a empresa → não gera parada na rota
        if fl_ret in ("sim", "s", "yes", "1", "true") and tipo_ret == "interna":
            if log_fn:
                log_fn(f"  → Pedido {ped.get('id')} ignorado (retirada interna).")
            continue

        id_ped = str(ped.get("id", "")).strip()
        id_cli = str(ped.get("cliente", "")).strip()

        # Buscar cliente por id ou nome
        cli_row = clientes[
            (clientes["id"].str.strip() == id_cli) |
            (clientes["nome"].str.strip().str.lower() == id_cli.lower())
        ]
        if cli_row.empty:
            if log_fn:
                log_fn(f"  ⚠ Cliente '{id_cli}' não encontrado. Pedido {id_ped} ignorado.")
            continue
        cli = cli_row.iloc[0]

        # Buscar horário do cliente para o dia da semana
        hor_row = horarios[
            (horarios["id_cliente"].str.strip() == str(cli.get("id", "")).strip()) &
            (horarios["dia_semana"].str.strip().str.lower() == dia_semana.lower())
        ]
        if hor_row.empty:
            if log_fn:
                log_fn(f"  ⚠ {cli['nome']} não atende {dia_semana}. Pedido {id_ped} ignorado.")
            continue
        hor = hor_row.iloc[0]

        # Converter horários — aceita fração decimal do Excel ou HH:MM
        h_entrada_min = hora_para_minutos(hor.get("hora_entrada", "08:00"))
        h_saida_min   = hora_para_minutos(hor.get("hora_saida",   "17:00"))

        tempo_est_raw = str(ped.get("tempo_estimado_servico", "")).strip()
        try:
            tempo_est = int(tempo_est_raw) if tempo_est_raw else 0
        except Exception:
            tempo_est = 0
        if tempo_est == 0:
            tipo_lower = str(ped.get("tipo_servico", "")).strip().lower()
            tempo_est = TEMPO_PADRAO_POR_TIPO.get(tipo_lower, TEMPO_PADRAO_GERAL)

        is_retirada_ext = (
            fl_ret in ("sim", "s", "yes", "1", "true") and tipo_ret == "externa"
        )

        paradas.append({
            "pedido_id":    id_ped,
            "cliente_nome": str(cli.get("nome", "")).strip(),
            "cep":          str(cli.get("cep", "")).strip(),
            "endereco":     str(cli.get("endereço", cli.get("endereco", ""))).strip(),
            "numero":       str(cli.get("numero_endereco", "")).strip(),
            "tipo":         str(ped.get("tipo_servico", "")).strip(),
            "prioridade":   str(ped.get("prioridade", "")).strip().lower(),
            "nf":           str(ped.get("nf", "")).strip(),
            "canhoto":      str(ped.get("canhoto", "")).strip(),
            "descricao":    str(ped.get("descricao", "")).strip(),
            "hora_pref":    str(ped.get("hora_preferencial", "")).strip(),
            "tempo_est":    tempo_est,
            "h_entrada":    minutos_para_hora(h_entrada_min),
            "h_saida":      minutos_para_hora(h_saida_min),
            "retirada_ext": is_retirada_ext,
            "lat":          None,
            "lon":          None,
            "dist_anterior_km": 0.0,
        })

    if not paradas:
        return [], None, "Nenhuma parada válida encontrada para este dia."

    # ── 3. Coordenadas ───────────────────────────────────────
    if log_fn:
        log_fn("Buscando coordenadas dos CEPs...")

    ceps_unicos = list({p["cep"] for p in paradas})
    coords = {}
    for cep in ceps_unicos:
        coords[cep] = buscar_coordenadas(cep, log_fn)
    for p in paradas:

        p["lat"], p["lon"] = coords.get(
            p["cep"],
            (None, None)
        )

        # CEP inválido
        if p["lat"] is None or p["lon"] is None:

            alerta_atual = p.get("alertas", "")

            if alerta_atual:
                alerta_atual += " | "

            p["alertas"] = alerta_atual + "CEP não encontrado"

    orig_lat, orig_lon = buscar_coordenadas(CEP_ORIGEM, log_fn)
    if not orig_lat:
        orig_lat, orig_lon = -23.1, -47.2

    # ── 4. Montar rota por score ─────────────────────────────
    # montar_rota considera em cada passo:
    #   urgentes    → bônus negativo enorme (sobem)
    #   retiradas   → penalidade máxima (ficam por último)
    #   abre tarde  → espera penaliza o score (fica para depois)
    #   fecha cedo  → atraso penaliza ainda mais (sobe na fila)
    if log_fn:
        log_fn("Calculando scores e montando rota...")

    rota = montar_rota(orig_lat, orig_lon, hora_para_minutos(jornada_entrada), paradas)
    

    # ── 6. Horários previstos e alertas ─────────────────────
    hora_cur_min  = hora_para_minutos(jornada_entrada)
    jornada_fim_m = hora_para_minutos(jornada_saida)
    
    almoco_ini_min = hora_para_minutos(almoco_inicio)
    almoco_fim_min = hora_para_minutos(almoco_fim)

    almoco_aplicado = False

    for i, p in enumerate(rota):
        dist           = p.get("dist_anterior_km", 0)
        mins_viagem    = tempo_viagem_min(dist)
        buf            = BUFFER_ENTRE_PARADAS if i > 0 else 0
        chegada_min    = hora_cur_min + mins_viagem + buf
        abertura_min   = hora_para_minutos(p["h_entrada"])
        fechamento_min = hora_para_minutos(p["h_saida"])

        # Respeitar hora preferencial
        if p["hora_pref"]:
            pref_min = hora_para_minutos(p["hora_pref"])
            if pref_min > 0 and chegada_min < pref_min:
                chegada_min = pref_min

        # Chegada efetiva: se chegou antes da abertura, aguarda
        chegada_efetiva = max(chegada_min, abertura_min)
        saida_local_min = chegada_efetiva + p["tempo_est"]
        
        # ─────────────────────────────
        # AJUSTE DE ALMOÇO
        # ─────────────────────────────

        if not almoco_aplicado:

            atravessou_almoco = (
                chegada_efetiva <= almoco_ini_min
                and saida_local_min >= almoco_ini_min
            )

            if atravessou_almoco:

                tempo_almoco = (
                    almoco_fim_min - almoco_ini_min
                )

                saida_local_min += tempo_almoco

                almoco_aplicado = True

                if log_fn:
                    log_fn(
                        f"Almoço aplicado após "
                        f"{p['cliente_nome']}")
        

        alertas = []
        # Mantém alertas anteriores
        if p.get("alertas"):
            alertas.append(p["alertas"])
        if chegada_min > fechamento_min:
            alertas.append(f"Fora da janela (fecha {p['h_saida']})")
        elif chegada_min < abertura_min:
            alertas.append(
                f"Aguarda abertura — chega {minutos_para_hora(chegada_min)}, "
                f"abre {p['h_entrada']}"
            )
        if saida_local_min > jornada_fim_m:
            alertas.append(f"Estoura jornada ({jornada_saida})")

        p["ordem"]                  = i + 1
        p["hora_chegada"]           = minutos_para_hora(chegada_efetiva)
        p["hora_saida_local"]       = minutos_para_hora(saida_local_min)
        p["alertas"]                = " | ".join(alertas)
        p["tempo_espera_min"]       = max(0, abertura_min - chegada_min)
        p["tempo_deslocamento_min"] = round(mins_viagem)
        p["tempo_servico_min"]      = p["tempo_est"]
        p["km_acumulado"]           = round(
            sum(r.get("dist_anterior_km", 0) for r in rota[:i+1]), 2)

        # Próxima parada parte da saída efetiva deste cliente
        hora_cur_min = saida_local_min

    # ── 7. Resumo de jornada ─────────────────────────────────
    usado_min = hora_cur_min - hora_para_minutos(jornada_entrada)
    jornada_info = {
        "entrada":        jornada_entrada,
        "saida":          jornada_saida,
        "saida_prevista": rota[-1]["hora_saida_local"] if rota else jornada_entrada,
        "total_min":      jornada_total,
        "usado_min":      usado_min,
        "estourou":       hora_cur_min > jornada_fim_m,
    }

    return rota, jornada_info, None

# ─────────────────────────────────────────────────────────────
#  SALVAR NA ABA ITINERARIO
# ─────────────────────────────────────────────────────────────

CABECALHOS = [
    "pedido_id", "data", "funcionario", "ordem",
    "cliente", "cep", "endereco",
    "tipo_servico", "nf", "canhoto", "descricao",
    "prioridade", "retirada_externa",
    "janela_entrada", "janela_saida",
    "hora_chegada", "hora_saida_local",
    "dist_anterior_km", "km_acumulado",
    "tempo_deslocamento_min", "tempo_espera_min", "tempo_servico_min",
    "semana", "mes", "alertas",
]


def salvar_itinerario(rota, data, nome_func):
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
    except PermissionError:
        return False, "Feche o arquivo Excel e tente novamente."
    except Exception as e:
        return False, str(e)

    # Criar aba se não existir
    if "Itinerario" not in wb.sheetnames:
        ws = wb.create_sheet("Itinerario")
        ws.append(CABECALHOS)
        for col in range(1, len(CABECALHOS) + 1):
            c = ws.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF", size=11)
            c.fill      = PatternFill("solid", fgColor="1A1916")
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        wb.save(ARQUIVO_EXCEL)
        wb = load_workbook(ARQUIVO_EXCEL)

    ws = wb["Itinerario"]

    # Apagar linhas cujo pedido_id está na nova rota
    ids_novos = {p["pedido_id"] for p in rota}
    linhas_apagar = []
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip() in ids_novos:
            linhas_apagar.append(row[0].row)
    for row_num in sorted(linhas_apagar, reverse=True):
        ws.delete_rows(row_num)

    # Estilos
    borda = Border(
        left=Side(style="thin",   color="E2E0D8"),
        right=Side(style="thin",  color="E2E0D8"),
        top=Side(style="thin",    color="E2E0D8"),
        bottom=Side(style="thin", color="E2E0D8"),
    )

    for p in rota:
        if p["prioridade"] == "urgente":
            cor = "FBEAEA"
        elif p["retirada_ext"]:
            cor = "EAF3DE"
        elif p["alertas"]:
            cor = "FAEEDA"
        else:
            cor = "FFFFFF"

        fill = PatternFill("solid", fgColor=cor)
        dt     = pd.to_datetime(data)
        semana = int(dt.isocalendar().week)
        mes    = dt.strftime("%Y-%m")
        linha = [
            p["pedido_id"], data, nome_func, p["ordem"],
            p["cliente_nome"], p["cep"],
            f"{p['endereco']}, {p['numero']}",
            p["tipo"], p["nf"], p.get("canhoto", ""), p["descricao"],
            p["prioridade"],
            "Sim" if p["retirada_ext"] else "Não",
            p["h_entrada"], p["h_saida"],
            p["hora_chegada"], p["hora_saida_local"],
            p["dist_anterior_km"], p.get("km_acumulado", 0),
            p.get("tempo_deslocamento_min", 0),
            p.get("tempo_espera_min", 0),
            p.get("tempo_servico_min", 0),
            semana, mes, p["alertas"],
        ]
        ws.append(linha)
        row_num = ws.max_row
        for col in range(1, len(linha) + 1):
            c = ws.cell(row=row_num, column=col)
            c.fill      = fill
            c.border    = borda
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_num].height = 30

    larguras = [12, 12, 16, 7, 28, 12, 34, 14, 14, 36, 12, 10, 12, 12, 12, 14, 10, 30]
    for i, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    # Registrar no histórico
    _salvar_historico(wb, data, nome_func, rota)

    try:
        wb.save(ARQUIVO_EXCEL)
        return True, "Salvo com sucesso."
    except PermissionError:
        return False, "Feche o arquivo Excel e tente novamente."
    except Exception as e:
        return False, str(e)


def _salvar_historico(wb, data, nome_func, rota):
    """Registra cada geração na aba Historico (útil para Power BI)."""
    if "Historico" not in wb.sheetnames:
        wh = wb.create_sheet("Historico")
        cab = ["data_geracao", "data_itinerario", "funcionario",
               "qtd_paradas", "km_total", "saida_prevista"]
        wh.append(cab)
        for col in range(1, len(cab) + 1):
            c = wh.cell(row=1, column=col)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="1A1916")
            c.alignment = Alignment(horizontal="center")
    else:
        wh = wb["Historico"]
    km_total   = round(sum(p.get("dist_anterior_km", 0) for p in rota), 1)
    saida_prev = rota[-1]["hora_saida_local"] if rota else ""
    wh.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        data, nome_func, len(rota), km_total, saida_prev,
    ])


def marcar_pedido_concluido(pedido_id):
    """Marca status do pedido como Concluído diretamente no Excel."""
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb["Pedidos"]
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        try:
            col_id     = headers.index("id") + 1
            col_status = headers.index("status") + 1
        except ValueError:
            return False, "Colunas 'id' ou 'status' não encontradas na aba Pedidos."
        for row in ws.iter_rows(min_row=2):
            if str(row[col_id - 1].value or "").strip() == str(pedido_id):
                row[col_status - 1].value = "Concluído"
                break
        wb.save(ARQUIVO_EXCEL)
        return True, "OK"
    except PermissionError:
        return False, "Feche o arquivo Excel e tente novamente."
    except Exception as e:
        return False, str(e)


#
#   PARA SALVAR O ITINARARIO EM PDF
#
def limpar_nome_arquivo(texto):
    texto = str(texto)

    caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']

    for c in caracteres_invalidos:
        texto = texto.replace(c, "-")

    return texto.replace(" ", "_")

def exportar_pdf(rota, jornada_info, nome_func, data):
    try:
        nome_arquivo = limpar_nome_arquivo(f"{data}_{nome_func}")

        caminho_pdf = os.path.join(
            PASTA_PDF,
            f"{nome_arquivo}.pdf"
        )

        doc = SimpleDocTemplate(
            caminho_pdf,
            pagesize=A4,
            rightMargin=1.2 * cm,
            leftMargin=1.2 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm,
        )

        elementos = []

        styles = getSampleStyleSheet()

        # ─────────────────────────────
        # TÍTULO
        # ─────────────────────────────

        titulo = Paragraph(
            f"<b>Itinerário — {nome_func}</b>",
            styles["Title"]
        )

        elementos.append(titulo)
        elementos.append(Spacer(1, 0.3 * cm))

        # ─────────────────────────────
        # RESUMO
        # ─────────────────────────────

        dist_total_pdf = sum(p.get("dist_anterior_km", 0) for p in rota)
        resumo = Paragraph(
            (
                f"<b>Data:</b> {data}  &nbsp;|&nbsp;  "
                f"<b>Jornada:</b> {jornada_info['entrada']} às {jornada_info['saida']}  &nbsp;|&nbsp;  "
                f"<b>Saída prevista:</b> {jornada_info['saida_prevista']}  &nbsp;|&nbsp;  "
                f"<b>Total:</b> {len(rota)} paradas / ~{dist_total_pdf:.0f} km"
            ),
            styles["BodyText"]
        )

        elementos.append(resumo)
        elementos.append(Spacer(1, 0.5 * cm))

        # ─────────────────────────────
        # TABELA
        # ─────────────────────────────

        dados = [[
            "#",
            "Cliente",
            # "Chegada",
            # "Saída",
            # "NF",
            "Tipo",
            "Endereço",
            "Alertas"
        ]]

        for p in rota:

            titulo_parada = Paragraph(
                f"<b>PARADA {p['ordem']} — {p['cliente_nome']}</b>",
                styles["Heading2"]
            )

            elementos.append(titulo_parada)
            elementos.append(Spacer(1, 0.2 * cm))

            endereco = (
                f"{p['endereco']}, {p['numero']}<br/>"
                f"CEP: {p['cep']}"
            )

            descricao = p["descricao"] or "Não informado"

            detalhes = f"""
            <b>📍 Endereço:</b><br/>
            {endereco}<br/><br/>
            
            <b>⏰ Horários:</b><br/>
            Chegada: {p['hora_chegada']}<br/>
            Saída prevista: {p['hora_saida_local']}<br/>
            Janela: {p['h_entrada']} às {p['h_saida']}<br/><br/>
            
            <b>🛠 Tipo de Serviço:</b>
            {p['tipo']}<br/><br/>

            <b>🧾 Descricao:</b><br/>
            {descricao}<br/><br/>

            """

            info = Paragraph(
                detalhes,
                styles["BodyText"]
            )

            elementos.append(info)
            elementos.append(Spacer(1, 0.7 * cm))

        # ─────────────────────────────
        # RODAPÉ
        # ─────────────────────────────

        dist_total = sum(
            p.get("dist_anterior_km", 0)
            for p in rota
        )

        elementos.append(Spacer(1, 0.5 * cm))

        rodape = Paragraph(
            (
                f"<b>Total de paradas:</b> {len(rota)}<br/>"
            ),
            styles["BodyText"]
        )

        elementos.append(rodape)

        # ─────────────────────────────
        # GERAR PDF
        # ─────────────────────────────

        doc.build(elementos)

        return True, caminho_pdf

    except Exception as e:
        return False, str(e)

# ─────────────────────────────────────────────────────────────
#  INTERFACE TKINTER
# ─────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Roteirizador — Itinerário")
        self.geometry("1400x750")
        self.minsize(900, 580)
        self.configure(bg="#F0EFEA")
        self.protocol("WM_DELETE_WINDOW", self.destroy)

        global ARQUIVO_EXCEL

        ARQUIVO_EXCEL = selecionar_planilha()

        if not ARQUIVO_EXCEL:

            messagebox.showwarning(
                "Atenção",
                "Nenhuma planilha selecionada."
            )

            self.destroy()
            return

        self.clientes, self.horarios, self.pedidos, self.funcionarios = ler_excel()
        self.rota_atual   = []
        self.jornada_info = None
        self._data_atual  = ""
        self._func_atual  = ""

        self._build_ui()
        self._popular_datas()

    # ── CONSTRUÇÃO DA UI ────────────────────────────────────

    def _build_ui(self):

        # Topo
        topo = tk.Frame(self, bg="#1A1916", height=52)
        topo.pack(fill="x")
        topo.pack_propagate(False)
        tk.Label(topo, text="Roteirizador",
                 font=("Helvetica", 15, "bold"),
                 bg="#1A1916", fg="#FFFFFF").pack(side="left", padx=18, pady=12)
        tk.Label(topo, text="Gerador de Itinerário",
                 font=("Helvetica", 10),
                 bg="#1A1916", fg="#888780").pack(side="left", pady=12)

        # Painel de seleção
        sel = tk.Frame(self, bg="#FFFFFF",
                       highlightthickness=1, highlightbackground="#E2E0D8")
        sel.pack(fill="x", padx=16, pady=(12, 0))
        inner = tk.Frame(sel, bg="#FFFFFF")
        inner.pack(fill="x", padx=14, pady=10)

        def lbl(texto, col):
            tk.Label(inner, text=texto, font=("Helvetica", 9),
                     bg="#FFFFFF", fg="#6B6860").grid(
                         row=0, column=col, sticky="w", padx=(0, 4), pady=(0, 2))

        lbl("Data", 0)
        self.cb_data = ttk.Combobox(inner, state="readonly", width=13,
                                    font=("Helvetica", 11))
        self.cb_data.grid(row=1, column=0, padx=(0, 14))
        self.cb_data.bind("<<ComboboxSelected>>", self._on_data_change)

        lbl("Funcionário", 1)
        self.cb_func = ttk.Combobox(inner, state="readonly", width=20,
                                    font=("Helvetica", 11))
        self.cb_func.grid(row=1, column=1, padx=(0, 14))

        lbl("Início da jornada", 2)
        self.ent_ini = ttk.Entry(inner, width=8, font=("Helvetica", 11))
        self.ent_ini.insert(0, "08:00")
        self.ent_ini.grid(row=1, column=2, padx=(0, 14))

        lbl("Fim da jornada", 3)
        self.ent_fim = ttk.Entry(inner, width=8, font=("Helvetica", 11))
        self.ent_fim.insert(0, "18:00")
        self.ent_fim.grid(row=1, column=3, padx=(0, 24))
        
        lbl("Almoço início", 4)
        self.ent_almoco_ini = ttk.Entry(
            inner,
            width=8,
            font=("Helvetica", 11)
        )
        self.ent_almoco_ini.insert(0, "12:00")
        self.ent_almoco_ini.grid(row=1, column=4, padx=(0, 14))


        lbl("Almoço fim", 5)
        self.ent_almoco_fim = ttk.Entry(
            inner,
            width=8,
            font=("Helvetica", 11)
        )
        self.ent_almoco_fim.insert(0, "13:00")
        self.ent_almoco_fim.grid(row=1, column=5, padx=(0, 24))
        # Linha exclusiva para botões
        frame_botoes = tk.Frame(inner, bg="#FFFFFF")

        frame_botoes.grid(
            row=2,
            column=0,
            columnspan=10,
            sticky="w",
            pady=(12, 0)
        )

        self.btn_calc = tk.Button(
            frame_botoes, text="Calcular rota",
            font=("Helvetica", 10, "bold"),
            bg="#1A6B3A", fg="#FFFFFF", activebackground="#145830",
            relief="flat", cursor="hand2", padx=14, pady=5,
            command=self._calcular,
        )
        self.btn_calc.grid(row=1, column=6, padx=(0, 8))

        self.btn_salvar = tk.Button(
            frame_botoes, text="Salvar no Excel",
            font=("Helvetica", 10, "bold"),
            bg="#0C447C", fg="#FFFFFF", activebackground="#083460",
            relief="flat", cursor="hand2", padx=14, pady=5,
            command=self._salvar, state="disabled",
        )
        self.btn_salvar.grid(row=1, column=7, padx=(0, 8))

        tk.Button(
            frame_botoes, text="↺ Recarregar",
            font=("Helvetica", 10),
            bg="#F0EFEA", fg="#444441",
            activebackground="#E2E0D8",
            relief="flat", cursor="hand2", padx=12, pady=5,
            command=self._recarregar,
        ).grid(row=1, column=8, padx=(0, 8))

        tk.Button(
            frame_botoes, text="Cancelar",
            font=("Helvetica", 10),
            bg="#F0EFEA", fg="#444441",
            activebackground="#E2E0D8",
            relief="flat", cursor="hand2", padx=12, pady=5,
            command=self._cancelar,
        ).grid(row=1, column=9)

        # Barra de jornada
        jf = tk.Frame(self, bg="#FFFFFF",
                      highlightthickness=1, highlightbackground="#E2E0D8")
        jf.pack(fill="x", padx=16, pady=(8, 0))
        ji = tk.Frame(jf, bg="#FFFFFF")
        ji.pack(fill="x", padx=14, pady=(7, 3))

        self.lbl_jornada = tk.Label(ji, text="Jornada: —",
                                    font=("Helvetica", 10),
                                    bg="#FFFFFF", fg="#6B6860")
        self.lbl_jornada.pack(side="left")

        self.lbl_pct = tk.Label(ji, text="",
                                font=("Helvetica", 10, "bold"),
                                bg="#FFFFFF", fg="#6B6860")
        self.lbl_pct.pack(side="right")

        barra_bg = tk.Frame(jf, bg="#E8E8E8", height=10)
        barra_bg.pack(fill="x", padx=14, pady=(0, 7))
        barra_bg.pack_propagate(False)
        self.barra = tk.Frame(barra_bg, bg="#1A6B3A", height=10)
        self.barra.place(relx=0, rely=0, relwidth=0, relheight=1)

        # Status
        self.lbl_status = tk.Label(self, text="",
                                   font=("Helvetica", 10),
                                   bg="#F0EFEA", fg="#6B6860", anchor="w")
        self.lbl_status.pack(fill="x", padx=18, pady=(6, 0))

        # Notebook: aba Rota + aba Log
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=16, pady=(8, 8))

        tab_rota = tk.Frame(self.nb, bg="#F0EFEA")
        self.nb.add(tab_rota, text="  Rota  ")

        tab_log = tk.Frame(self.nb, bg="#1A1916")
        self.nb.add(tab_log, text="  Log  ")
        self.txt_log = tk.Text(
            tab_log, font=("Courier", 10),
            bg="#1A1916", fg="#C8C5BC",
            state="disabled", wrap="word", relief="flat", bd=0,
        )
        sb_log_v = ttk.Scrollbar(tab_log, orient="vertical", command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sb_log_v.set)
        sb_log_v.pack(side="right", fill="y")
        self.txt_log.pack(fill="both", expand=True, padx=2, pady=2)

        tf = tab_rota

        cols = ("ordem", "cliente", "tipo", "nf",
                "janela", "chegada", "saida_local", "dist", "alertas")
        self.tree = ttk.Treeview(tf, columns=cols,
                                 show="headings", selectmode="browse")

        hdrs = {
            "ordem":       ("#",           50),
            "cliente":     ("Cliente",    220),
            "tipo":        ("Tipo",        90),
            "nf":          ("NF",          90),
            "janela":      ("Janela do Cliente",     120),
            "chegada":     ("Chegada",     75),
            "saida_local": ("Saída local", 80),
            "dist":        ("Dist. km",    70),
            "alertas":     ("Alertas",    220),
        }
        for col, (label, width) in hdrs.items():
            self.tree.heading(col, text=label)
            anc = "center" if col in ("ordem", "chegada", "saida_local", "dist") else "w"
            self.tree.column(col, width=width, anchor=anc)

        self.tree.tag_configure("urgente",   background="#FBEAEA", foreground="#7A1F1F")
        self.tree.tag_configure("retirada",  background="#EAF3DE", foreground="#27500A")
        self.tree.tag_configure("alerta",    background="#FAEEDA", foreground="#633806")
        self.tree.tag_configure("normal",    background="#FFFFFF", foreground="#1A1916")
        self.tree.tag_configure("concluido", background="#F0F0F0", foreground="#999999")

        # Menu de contexto (botão direito) — marcar como concluído
        self._ctx = tk.Menu(self, tearoff=0)
        self._ctx.add_command(label="✓  Marcar como Concluído",
                              command=self._marcar_concluido_selecionado)
        self.tree.bind("<Button-3>", self._show_ctx)

        sb_v = ttk.Scrollbar(tf, orient="vertical",   command=self.tree.yview)
        sb_h = ttk.Scrollbar(tf, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
        sb_v.pack(side="right",  fill="y")
        sb_h.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        # Rodapé
        rod = tk.Frame(self, bg="#FFFFFF",
                       highlightthickness=1, highlightbackground="#E2E0D8")
        rod.pack(fill="x", padx=16, pady=(0, 12))
        ir = tk.Frame(rod, bg="#FFFFFF")
        ir.pack(fill="x", padx=14, pady=8)

        self.s_total    = self._stat(ir, "Total de paradas",    "—", 0)
        self.s_dist     = self._stat(ir, "Distância estimada",  "—", 1)
        self.s_urgentes = self._stat(ir, "Urgentes",            "—", 2)
        self.s_retirada = self._stat(ir, "Retiradas externas",  "—", 3)
        
        self.btn_fechar = tk.Button(
                frame_botoes,
                text="Fechar",
                font=("Helvetica", 10, "bold"),
                bg="#A32D2D",
                fg="#FFFFFF",
                relief="flat",
                cursor="hand2",
                padx=14,
                pady=5,
                command=self.destroy
            )

        self.btn_fechar.grid(row=1, column=8, padx=(8, 0))

    def _stat(self, parent, titulo, valor, col):
        f = tk.Frame(parent, bg="#F1EFE8")
        f.grid(row=0, column=col, padx=(0, 10), ipadx=12, ipady=6, sticky="w")
        tk.Label(f, text=titulo, font=("Helvetica", 9),
                 bg="#F1EFE8", fg="#888780").pack(anchor="w", padx=8, pady=(4, 0))
        lbl = tk.Label(f, text=valor, font=("Helvetica", 18, "bold"),
                       bg="#F1EFE8", fg="#1A1916")
        lbl.pack(anchor="w", padx=8, pady=(0, 4))
        return lbl

    # ── FILTROS ─────────────────────────────────────────────

    def _popular_datas(self):
        status_exc = ["concluido", "concluído", "cancelado"]
        pend = self.pedidos[
            ~self.pedidos["status"].str.lower().str.strip().isin(status_exc)
        ]
        datas = sorted(pend["data_trabalho"].str.strip().unique().tolist())
        self.cb_data["values"] = datas
        if datas:
            self.cb_data.set(datas[0])
            self._on_data_change()

    def _on_data_change(self, event=None):
        data = self.cb_data.get().strip()
        status_exc = ["concluido", "concluído", "cancelado"]
        pend = self.pedidos[
            (self.pedidos["data_trabalho"].str.strip() == data) &
            (~self.pedidos["status"].str.lower().str.strip().isin(status_exc))
        ]
        funcs = sorted(pend["funcionario"].str.strip().unique().tolist())
        self.cb_func["values"] = funcs
        self.cb_func.set(funcs[0] if funcs else "")

    # ── CALCULAR ────────────────────────────────────────────

    def _calcular(self):
        data  = self.cb_data.get().strip()
        func  = self.cb_func.get().strip()
        j_ini = self.ent_ini.get().strip()
        j_fim = self.ent_fim.get().strip()
        almoco_ini = self.ent_almoco_ini.get().strip()
        almoco_fim = self.ent_almoco_fim.get().strip()

        if not data or not func:
            messagebox.showwarning("Atenção", "Selecione data e funcionário.")
            return

        for h in (j_ini, j_fim):
            partes = h.split(":")
            if len(partes) != 2 or not all(x.isdigit() for x in partes):
                messagebox.showwarning("Horário inválido",
                                       f"Use o formato HH:MM. Valor inválido: {h}")
                return

        if hora_para_minutos(j_ini) >= hora_para_minutos(j_fim):
            messagebox.showwarning("Atenção",
                                   "O início da jornada deve ser antes do fim.")
            return

        self.btn_salvar.config(state="disabled")
        self.rota_atual = []
        self._limpar_tabela()
        self.barra.place(relwidth=0)
        self.lbl_jornada.config(text="Jornada: —")
        self.lbl_pct.config(text="")
        self._set_status("Calculando rota...")
        self._log(f"\n── Iniciando: {func} | {data} ──")

        def _log_dual(msg):
            self._set_status(msg)
            self._log(msg)

        def run():
            rota, jornada, erro = calcular_rota(
                data, func, j_ini, j_fim, almoco_ini, almoco_fim,
                self.clientes, self.horarios, self.pedidos,
                log_fn=_log_dual,
            )
            self.after(0, lambda: self._exibir(rota, jornada, erro, data, func))

        threading.Thread(target=run, daemon=True).start()

    def _exibir(self, rota, jornada, erro, data, func):
        if erro:
            self._set_status(f"⚠ {erro}")
            return

        self.rota_atual   = rota
        self.jornada_info = jornada
        self._data_atual  = data
        self._func_atual  = func

        self._limpar_tabela()
        for p in rota:
            tag = ("urgente"  if p["prioridade"] == "urgente" else
                   "retirada" if p["retirada_ext"] else
                   "alerta"   if p["alertas"] else "normal")

            self.tree.insert("", "end", iid=p["pedido_id"], values=(
                p["ordem"],
                ("↩ " if p["retirada_ext"] else "") + p["cliente_nome"],
                p["tipo"],
                p["nf"] or "—",
                f"{p['h_entrada']} – {p['h_saida']}",
                p["hora_chegada"],
                p["hora_saida_local"],
                f"{p['dist_anterior_km']:.1f}",
                p["alertas"] or "✓",
            ), tags=(tag,))

        dist_total = sum(p.get("dist_anterior_km", 0) for p in rota)
        self.s_total.config(text=str(len(rota)))
        self.s_dist.config(text=f"~{dist_total:.0f} km")
        self.s_urgentes.config(text=str(sum(1 for p in rota if p["prioridade"] == "urgente")))
        self.s_retirada.config(text=str(sum(1 for p in rota if p["retirada_ext"])))

        if jornada:
            pct = min(jornada["usado_min"] / max(jornada["total_min"], 1), 1.0)
            cor = "#A32D2D" if jornada["estourou"] else \
                  "#854F0B" if pct > 0.85 else "#1A6B3A"
            self.barra.config(bg=cor)
            self.barra.place(relwidth=pct)
            self.lbl_jornada.config(
                text=(f"Jornada {jornada['entrada']} – {jornada['saida']}  |  "
                      f"Saída prevista: {jornada['saida_prevista']}")
            )
            txt = f"{pct * 100:.0f}% utilizado"
            if jornada["estourou"]:
                txt += "  ⚠ ESTOURA JORNADA"
            self.lbl_pct.config(text=txt,
                                 fg="#A32D2D" if jornada["estourou"] else "#6B6860")

        self.btn_salvar.config(state="normal")
        self._set_status(f"✓ {len(rota)} paradas | ~{dist_total:.0f} km")
        self._log(f"✓ Rota calculada: {len(rota)} paradas | ~{dist_total:.0f} km")

    # ── SALVAR ──────────────────────────────────────────────

    def _salvar(self):
        if not self.rota_atual:
            return
        if self.jornada_info and self.jornada_info["estourou"]:
            if not messagebox.askyesno(
                "⚠ Jornada estourada",
                "A rota ultrapassa o horário de saída do funcionário.\n"
                "Deseja salvar mesmo assim?"
            ):
                return

        self._set_status("Salvando no Excel...")
        self.btn_salvar.config(state="disabled")
        ok, msg = salvar_itinerario(self.rota_atual, self._data_atual, self._func_atual)

        if ok:
            self._set_status("✓ Salvo na aba 'Itinerario'.")
            self._log(f"✓ Excel salvo: {self._func_atual} | {self._data_atual}")

            pdf_ok, pdf_msg = exportar_pdf(
                self.rota_atual,
                self.jornada_info,
                self._func_atual,
                self._data_atual
            )

            self._set_status("✓ Salvo na aba 'Itinerario'.")

            mensagem = (
                f"Itinerário de {self._func_atual} "
                f"({self._data_atual}) salvo com sucesso."
            )

            if pdf_ok:
                mensagem += f"\n\nPDF gerado em:\n{pdf_msg}"
            else:
                mensagem += f"\n\nErro ao gerar PDF:\n{pdf_msg}"

            messagebox.showinfo("Salvo", mensagem)
        else:
            self._set_status(f"✗ Erro: {msg}")
            messagebox.showerror("Erro ao salvar", msg)

        self.btn_salvar.config(state="normal")

    # ── RECARREGAR ──────────────────────────────────────────

    def _recarregar(self):
        """Relê o Excel sem fechar o programa."""
        try:
            data_ant = self.cb_data.get()
            func_ant = self.cb_func.get()
            self.clientes, self.horarios, self.pedidos, self.funcionarios = ler_excel()
            self._popular_datas()
            if data_ant in list(self.cb_data["values"]):
                self.cb_data.set(data_ant)
                self._on_data_change()
                if func_ant in list(self.cb_func["values"]):
                    self.cb_func.set(func_ant)
            self._log("↺ Dados recarregados do Excel com sucesso.")
            self._set_status("↺ Recarregado.")
        except Exception as e:
            messagebox.showerror("Erro ao recarregar", str(e))

    # ── CANCELAR ────────────────────────────────────────────

    def _cancelar(self):
        """Limpa a rota exibida sem fechar o programa."""
        self._limpar_tabela()
        self.rota_atual   = []
        self.jornada_info = None
        self.barra.place(relwidth=0)
        self.lbl_jornada.config(text="Jornada: —")
        self.lbl_pct.config(text="")
        self.btn_salvar.config(state="disabled")
        for s in (self.s_total, self.s_dist, self.s_urgentes, self.s_retirada):
            s.config(text="—")
        self._set_status("Cancelado.")

    # ── MARCAR CONCLUÍDO ────────────────────────────────────

    def _show_ctx(self, event):
        row = self.tree.identify_row(event.y)
        if row:
            self.tree.selection_set(row)
            self._ctx.post(event.x_root, event.y_root)

    def _marcar_concluido_selecionado(self):
        sel = self.tree.selection()
        if not sel:
            return
        pedido_id = sel[0]
        nome_cli  = self.tree.item(pedido_id, "values")[1]
        if not messagebox.askyesno("Confirmar",
                f"Marcar '{nome_cli}' como Concluído no Excel?"):
            return
        ok, msg = marcar_pedido_concluido(pedido_id)
        if ok:
            self.tree.item(pedido_id, tags=("concluido",))
            self._log(f"✓ Concluído: pedido {pedido_id} — {nome_cli}")
            self._set_status(f"✓ Pedido {pedido_id} marcado como Concluído.")
        else:
            messagebox.showerror("Erro", msg)

    # ── LOG ─────────────────────────────────────────────────

    def _log(self, msg):
        """Adiciona linha com timestamp na aba Log."""
        ts    = datetime.now().strftime("%H:%M:%S")
        linha = "[" + ts + "] " + str(msg) + "\n"
        try:
            self.txt_log.config(state="normal")
            self.txt_log.insert("end", linha)
            self.txt_log.see("end")
            self.txt_log.config(state="disabled")
        except Exception:
            pass

    # ── UTILS ───────────────────────────────────────────────

    def _limpar_tabela(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _set_status(self, msg):
        self.lbl_status.config(text=msg)
        self.update_idletasks()


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
    